# 清除终端历史命令：shift+alt+c
# 同时选中编辑相同内容：ctrl+shift+l

import pandas as pd
import numpy as np
import os
import datetime
from openpyxl import load_workbook

#####################  230523  #############################

# 添加dataframe类型数据至新建sheet页

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

#load_workbook将excel文件加载至内存中，对其数据进行：读取/修改单元格内容、创建新的sheet页
load_workbook

da_test=pd.DataFrame([['one','two'],['three','four']],columns=['hi','no'])

book = load_workbook('D:/市场看板/市场/1.转化数据汇总/test.xlsx')

sheet_cu = book.active
sheet_cu['A1'].value  #读取A1单元格内容

new_sheet = book.create_sheet(title='new')  #创建sheet页

# da_test.to_excel('D:/市场看板/市场/事业单&公务员位转化数据.xlsx',sheet_name=week)

rows = dataframe_to_rows(da_test,index=False,header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        new_sheet.cell(row=r_idx, column=c_idx, value=value)

book.save('D:/市场看板/市场/1.转化数据汇总/test.xlsx')



# 将结果追加写入已有excel
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

book = load_workbook('D:/项目/事业单位-联报单科(YH项目看板).xlsx')
sheet = book.active
rows = dataframe_to_rows(da_test,index=False,header=False)

for r_idx, row in enumerate(rows, sheet.max_row + 1):
    for c_idx, value in enumerate(row, 1):
        sheet.cell(row=r_idx, column=c_idx, value=value)




# 添加数组类型数据至已有sheet页
import openpyxl
workbook = openpyxl.load_workbook('D:/市场看板/市场/1.转化数据汇总/0515期_事业单&公务员位转化数据_汇总.xlsx')
sheet = workbook.active

datas = [['lucas','28','m'],['emma','22','f']]   
for data in datas:
    sheet.append(data)

workbook.save('D:/市场看板/市场/1.转化数据汇总/0515期_事业单&公务员位转化数据_汇总.xlsx')



# 添加dataframe类型数据至已有sheet页
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
workbook = load_workbook('D:/市场看板/市场/1.转化数据汇总/test.xlsx')
sheet = workbook.active

new_data=pd.DataFrame({'name':['test','test','test']})
rows = dataframe_to_rows(new_data,index=False,header=True)

for r_idx, row in enumerate(rows, sheet.max_row + 1):
    for c_idx, value in enumerate(row, 1):
        sheet.cell(row=r_idx, column=c_idx, value=value)

workbook.save('D:/市场看板/市场/1.转化数据汇总/test.xlsx')




##################################################


# 1.读入读出excel
data1=pd.read_excel(r'D:\服务相关\数据源\6.每月绩效\data_gwy.xlsx', dtype={'用户编号':'str'}) 
#读入数据 names设置列名,usecols选择需要的列,nrows设置读入文件行数,parse_dates将某列指定为时间类型,iterator=True，分块读入数据，df.get_chunk(n)

data1.dtypes


# 分组汇总
data2=data1.groupby(by=['用户编号'],as_index=False).agg({'看课进度10%节数':'sum','课节总数(正式课节+赠课课节)':'sum'})   
data2.columns


# 数据透视表
data2_2 = pd.pivot_table(data1,values=['正式课节总数','正式课节看课进度10%节数'],index='用户编号',aggfunc=np.sum,fill_value=0)
data2_2.reset_index(drop=False)


cols_watch=['用户id','看课节数','总课节数']
data2.columns=cols_watch     #变量重命名
data2.dtypes



# 读出到同一excel不同sheet页
#### 方法1
writer = pd.ExcelWriter(r'D:\代码\test\test_watch.xlsx')
data1.to_excel(writer, sheet_name='data', na_rep='', index=False)
data2.to_excel(writer, sheet_name='学员完课汇总', index=False)
writer.close()

#### 方法2
with pd.ExcelWriter(r'D:\代码\test\test_watch2.xlsx') as writer2:
    data1.to_excel(writer2, sheet_name='data', na_rep='', index=False)
    data2.to_excel(writer2, sheet_name='学员完课汇总', index=False)




# 向已有excel中追加数据##############################################
os.chdir('D:\\代码\\test')

s1 = pd.DataFrame(np.array([['s1', 's1', 's1', 's1']]), columns=['a', 'b', 'c', 'd'])
s2 = pd.DataFrame(np.array([['s2', 's2', 's2', 's2']]), columns=['a', 'b', 'c', 'd'])
# s4 只有3列,并且列顺序被打乱,以模拟新数据与元数据的差异
s4 = pd.DataFrame(np.array([['s4b', 's4d', 's4c']]), columns=['b', 'd', 'c'])

with pd.ExcelWriter("test_append.xlsx") as writer:
    # 先写入两个sheet
    s1.to_excel(writer, sheet_name="111", index=False)
    s2.to_excel(writer, sheet_name="222", index=False)


# 读入excel
df = pd.read_excel('test_append.xlsx', sheet_name='111')
row = df.shape[0]	# 获取原数据的行数

s4 = pd.concat([pd.DataFrame(columns=df.columns), s4], ignore_index=True)  # 将新数据格式化成原数据的模样,以解决数据列之间的差异
s4


# 将s4数据写入excel的第一个工作表中

book = load_workbook("test_append.xlsx")      #加载excel文件

book.active               #获取当前活动的工作表,默认最后关闭时的sheet
book.worksheets           #获取全部工作表
book.get_sheet_names()    #获取全部sheet名

writer = pd.ExcelWriter("D:\\代码\\test\\test_append.xlsx", engine='openpyxl')

writer = pd.ExcelWriter("D:\\代码\\test\\test_append.xlsx")
writer.book = book
writer.sheets = {sheet.title: sheet for sheet in book.worksheets}
s4.to_excel(writer, sheet_name='111', startrow=row + 1, index=False, header=False)
writer.close()

with pd.ExcelWriter("test_append.xlsx",engine='openxl') as writer:
    writer.book = book
    writer.sheets = {sheet.title: sheet for sheet in book.worksheets}
    # 追加新数据,追加前必须先格式化新数据,否则新数据缺少列,或是列顺序不对会导致数据紊乱
    s4.to_excel(writer, sheet_name='111', startrow=row + 1, index=False, header=False)

###############################################




# 2.批量处理excel
# 获取文件集合
path='D:\\代码\\test\\功底测test\\' #批量处理excel所在的文件夹路径
filename = os.listdir(path)   #获取该文件夹目录下的文件
filename


# 创建循环，遍历读入每一个Excel文件
for i in range(len(filename)):
    df = pd.read_excel(path+filename[i])
    df.dropna(how='all',axis=1,inplace=True)
    bins=[0,89,200]
    df['if_pass'] = pd.cut(df['8月功底测分数'],bins,labels=['不及格','及格'],include_lowest=True)  #数据分类
    df.to_excel(path + filename[i].split('.')[0] +'-文件处理成功.xlsx',index=False)  # 读出数据&截取字符
    print(filename[i].split('.')[0]+'************文件处理成功')



# root获取所有文件夹路径
for root,dirs,files in os.walk('D:\\服务相关\\数据源\\2.2NPS名单'):
    print(root)
    for file in files:
        print(file)


# 2.1批量合并同一文件夹下的所有excel文件
filelist=[]
for root,dirs,files in os.walk('D:\\服务相关\\数据源\\2.NPS'):   
    for file in files:
        if file.endswith('xlsx'):
            filelist.append(os.path.join(root,file))
filelist


writer3=pd.ExcelWriter('D:\\代码\\test\\test_write.xlsx')
for file in filelist:
    dfs=pd.read_excel(file)
    dfs.to_excel(writer3,sheet_name=file.split('\\')[-1].split('.')[0],index=False)
writer3.close()




# 3.调整excel格式 ################################# 
data2.head()
data2.shape


def modify_excel(writer,data):
    workbook = writer.book
    # fmt = workbook.add_format({"font_name": u"微软雅黑"})  #字体
    col_fmt = workbook.add_format(
        {'bold': True, 'font_name': u'微软雅黑', 'font_size': 11, 'valign': 'vcenter', 'align': 'left'})
    detail_fmt = workbook.add_format(
        {"font_name": u"微软雅黑", 'font_size': 10, 'valign': 'vcenter', 'align': 'left'})
    worksheet1 = writer.sheets['完课汇总']
    
    for col_num, value in enumerate(data.columns.values):
        worksheet1.write(0, col_num, value, col_fmt)
    worksheet1.set_column('A:C',20)
    worksheet1.set_row(0,21)
    for i in range(1, len(data)+1):
        worksheet1.set_row(i, 20, detail_fmt)
    

with pd.ExcelWriter(r'D:\代码\test\test_modify.xlsx', engine="xlsxwriter") as writer1:
    data2.loc[:100].to_excel(writer1, sheet_name='完课汇总', index=False)
    modify_excel(writer1,data2)



# 提取sheet页的全部标题
for col_num, value in enumerate(data2.columns.values):
    print(col_num, value)


'''
https://blog.csdn.net/qq_45219614/article/details/126002629  设置格式
'''



# 4.pandas数据处理
da1=pd.read_excel(r'D:\代码\test\test.xlsx')
da1



# 1.筛选行列
## 多条件筛选
Table = pd.DataFrame({'date': ['2019/6/1', '2019/7/2', '2019/6/6', '2019/6/17', '2019/7/4', '2019/6/13', '2019/6/14', '2019/6/21', '2019/6/17'], \
    'order_id': [i+1 for i in range(9)],
    'commodity_code': ['S1', 'S2', 'S3', 'S5', 'S5', 'S2', 'S9', 'S11', 'S9'], \
    'commodity_name': ['标准美式','瑞纳冰', '加浓美式', '拿铁', '拿铁', '瑞纳冰', '菠萝卷', '坚果', '菠萝卷'], \
    'category_name': ['饮品', '饮品', '饮品', '饮品', '饮品', '饮品', '食品', '食品', '食品']})

Table.columns.to_list()

#过滤出购买超过一单的商品对应的所有订单信息
temp = Table.groupby('commodity_code')['commodity_code'].count()  #分组计数法
temp[temp>1].index

Table[Table['commodity_code'].isin(temp[temp>1].index)]


# 4.2.pandas的value_counts()函数法
multi = Table['commodity_name'].value_counts()>1
type(multi)

multi[multi].index


Table[Table['commodity_name'].isin(multi[multi].index)]




# 4.3.删除不需要的行列
da_drop = pd.DataFrame(np.arange(12).reshape(3, 4),columns=['A', 'B', 'C', 'D'])
da_drop

da_drop.drop(["C","A"], axis=1, inplace=False)   #删除多列
da_drop.drop(0)  #删除行，axis默认是0
 

# 4.4.删除重复值
da4 = da1[['paid_date','user_type']].drop_duplicates()
da4.shape


# 4.5删除有NaN值的行列
import numpy as np 
da = pd.DataFrame({"name": ['Alfred', 'Batman', 'Catwoman'],\
    "toy": [np.nan, 'Batmobile', 'Bullwhip'],\
    "born": [pd.NaT, pd.Timestamp("1940-04-25"), pd.NaT]})
da

da.dropna(subset=["toy"])  #默认how是any，只要有一个na则会删除此列
da.dropna(subset=["born"])
da.dropna(subset=["toy","born"],how="all")




# 4.6.重命名
da1.columns

da5=da1.rename(columns={"final_paid_timestamp":"paid_timestamp","user_type":"type"})
da5.head()

# 直接修改数据框的列名
cols=['paid_date','paid_time','user_id','type']
da5.columns=cols   #会直接修改数据框
da5.head()

# 也可以在读入数据集时对变量重命名



# 4.7.改变列顺序
da1
pd.DataFrame(da1,columns=['user_type','user_number','paid_date'])


# 4.1当只需改动一个列的顺序时，可以先drop该列再insert到指定位置
da1_new=pd.DataFrame(da1,columns=['user_type'])
da1_new

da1_change = da1.drop('user_type',axis=1)
da1_change

da1_change.insert(0,'user_type',da1_new)
da1_change



# 4.8.日期处理相关
dt1 = pd.to_datetime('2022-05-10')
dt1.month         #只有一个日期不需要用.dt

dts=pd.DataFrame([['2023-03-25'],['2024-05-10']])
dts

# 字符串转日期
dts[0]=pd.to_datetime(dts[0])
dts[0].dt.month     #多个日期时需要用dt


# 日期转字符串
dt1.strftime('%Y-%m-%d')     
dts[0].strftime('%Y-%m-%d') 
dts.dt.strftime('%Y-%m')

t = datetime.datetime.now().date()-datetime.timedelta(days=1)  # 昨日


# 4.9.分组统计group by 
da_gb = pd.DataFrame({'Animal': ['Falcon', 'Falcon', 'Parrot', 'Parrot'], 'Max Speed': [380., 370., 24., 26.]})
da_gb

col_gb=da_gb['Animal']

da_gb.groupby(col_gb, as_index=False).max()  #分组字段不会被设为索引，默认是会的
da_gb.groupby(['Animal'], as_index=False).max()


# df.set_index(keys)--将keys列指定为索引；reset_index()用来还原索引
da_gb.reset_index(drop=False)  #会保留索引列

da1.groupby(['user_type']).agg({'user_number':'unique'}).reset_index()  #返回一个去重的列表
da1.groupby(['user_type']).agg({'user_number':'nunique'}).reset_index()  #分组去重计数, 更常用

da1.groupby(['user_type']).agg({"user_number":["unique","nunique"]})     #对一个字段实现两种计算方式



# transform 会传到每行数据上，agg是会返回一个聚合结果(transform在原数据基础上新加一列)
da1.groupby(['user_type'])["user_number"].transform("nunique")       #用来给原始数据框添加统计列
da1['user_cnt']=da1.groupby(['user_type'])["user_number"].transform("nunique")
da1




# 7.数据框合并 
df11=pd.DataFrame({"A":np.arange(10)})
df11

df22=pd.DataFrame([7,8],columns=["A"])
df22

pd.concat([df11,df22])   #pd.concat只是数据框的拼接
pd.concat([df11,df22],ignore_index=True)  #会更新index



df1 = pd.DataFrame([[1,2,3],['a','b','c']],columns=['T','Y','U'])
df2 = pd.DataFrame([[0,2,4],['a','b','m']],columns=['T','Y','I'])


# 将两个数据框根据行轴/列轴拼接在一起
pd.concat([df1,df2])
pd.concat([df1,df2],axis=1)  


# 将两个数据库根据指定列合并
df1.merge(df2, on=['T','Y'])  # 默认inner合并
df1.merge(df2, on=['T','Y'],how='outer')


# 同merge
df1.join(df2.set_index(['T','Y']),on=['T','Y'],how='inner')
df1.join(df2.set_index(['T','Y']),on=['T','Y'],rsuffix='_df2')




# 5.正则
import re
data1['assi']=1
data_re = data1.pivot_table(index=['班级编号','班级名称'],aggfunc={'assi':sum}).reset_index(drop=False)

pattern = re.compile(r"[\u4e00-\u9fa5]")

pattern.findall(data_re['班级名称'][0])



date_va = '2023-04.*'
re_da['喜报日期'].str.contains(date_va)


# 剔除不通过记录
re_da['审核结果'].value_counts().sort_values(ascending=False)

re_da2 = re_da['喜报日期'].dropna()

fail = '不通过|重复|缺少'

re_da2[re_da2.str.contains(fail)].index
